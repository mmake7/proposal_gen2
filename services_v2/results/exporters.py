"""v2 결과 → Excel / Markdown 출력.

v1 excel_exporter는 v2 RfpAnalysisResult.to_dict()와 호환이므로 재사용.
추가로 v2 전용 메타 시트(classification/validation/agent_trace) 부착.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.excel_exporter import (
    export_to_excel as _v1_export_to_excel,
    _BODY_ALIGNMENT, _BODY_FONT, _HEADER_ALIGNMENT, _HEADER_FILL,
    _HEADER_FONT, _LABEL_FILL, _LABEL_FONT, _THIN_BORDER, _TITLE_FONT,
    _auto_column_width,
)
from services_v2.results.rfp_analysis import RfpAnalysisResult


# ── Excel ──────────────────────────────────────────────────

def to_excel(result: RfpAnalysisResult) -> io.BytesIO:
    """v2 결과 → Excel BytesIO.

    시트: 사업개요/요구사항/배점기준/제안목차 (v1 호환) + 분석메타/검증결과/에이전트로그
    """
    base_dict = result.to_dict()
    buf = _v1_export_to_excel(base_dict)
    buf.seek(0)
    wb = load_workbook(buf)

    _build_classification_sheet(wb, result)
    _build_validation_sheet(wb, result)
    _build_estimation_sheets(wb, result)
    _build_agent_trace_sheet(wb, result)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_classification_sheet(wb: Workbook, result: RfpAnalysisResult) -> None:
    ws = wb.create_sheet(title='분석메타')
    ws.merge_cells('A1:B1')
    ws['A1'].value = 'RFP 분류 (v2 Classifier)'
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    cls = result.classification
    if cls is None:
        ws['A3'].value = '분류 정보 없음'
        ws['A3'].font = _BODY_FONT
        return

    labels = [
        ('org_type', '발주기관 유형'),
        ('req_id_scheme', '요구사항 ID 체계'),
        ('req_format', '요구사항 형식'),
        ('scoring_format', '배점 형식'),
        ('toc_location', '목차 위치'),
        ('doc_size_tier', '문서 규모'),
        ('language_dialect', '어휘 특이점'),
        ('notes', '분류 메모'),
    ]
    row = 3
    for key, label in labels:
        value = getattr(cls, key, '') or '-'
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _LABEL_FONT
        lc.fill = _LABEL_FILL
        lc.border = _THIN_BORDER
        lc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _BODY_FONT
        vc.border = _THIN_BORDER
        vc.alignment = _BODY_ALIGNMENT
        ws.row_dimensions[row].height = max(28, min(80, 20 + len(str(value)) // 4))
        row += 1
    _auto_column_width(ws, [20, 80])


def _build_validation_sheet(wb: Workbook, result: RfpAnalysisResult) -> None:
    ws = wb.create_sheet(title='검증결과')
    ws.merge_cells('A1:C1')
    ws['A1'].value = '검증 결과 (Validator)'
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    val = result.validation
    if val is None:
        ws['A3'].value = '검증 정보 없음'
        ws['A3'].font = _BODY_FONT
        return

    ws['A3'].value = '신뢰도'
    ws['A3'].font = _LABEL_FONT
    ws['A3'].fill = _LABEL_FILL
    ws['B3'].value = f'{val.confidence:.2f}'
    ws['B3'].font = Font(name='맑은 고딕', bold=True, size=14,
                        color='2B579A' if val.confidence >= 0.7 else 'C0504D')
    ws.merge_cells('B3:C3')

    headers = ['심각도', '필드', '메시지']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGNMENT
        c.border = _THIN_BORDER

    row = 6
    severity_colors = {
        'error': 'F2DCDB', 'warning': 'FFF2CC', 'info': 'E8EEF7',
    }
    for issue in val.issues:
        severity = issue.get('severity', 'info')
        ws.cell(row=row, column=1, value=severity).font = _BODY_FONT
        ws.cell(row=row, column=2, value=issue.get('field', '')).font = _BODY_FONT
        ws.cell(row=row, column=3, value=issue.get('message', '')).font = _BODY_FONT
        fill = PatternFill(start_color=severity_colors.get(severity, 'FFFFFF'),
                          end_color=severity_colors.get(severity, 'FFFFFF'),
                          fill_type='solid')
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT
        row += 1

    _auto_column_width(ws, [10, 15, 80])


def _build_estimation_sheets(wb: Workbook, result: RfpAnalysisResult) -> None:
    """영업 견적 — WBS / 조직도 / M/M 3시트."""
    est = result.estimation
    if est is None:
        return

    # 1) WBS 시트
    ws = wb.create_sheet(title='WBS')
    ws.merge_cells('A1:E1')
    ws['A1'].value = f'WBS (사업기간 {est.project_months:.0f}개월)'
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36
    headers = ['단계', '주차', '기간(주)', '주요 활동', '산출물']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGNMENT
        c.border = _THIN_BORDER
    row = 4
    for phase in est.wbs:
        week_range = (f'{phase.week_start}~{phase.week_end}'
                     if phase.week_start else '-')
        values = [
            phase.name, week_range, phase.duration_weeks,
            '\n'.join(f'• {a}' for a in phase.activities),
            '\n'.join(f'• {d}' for d in phase.deliverables),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT
        ws.row_dimensions[row].height = max(
            28, 18 * max(len(phase.activities), len(phase.deliverables)))
        row += 1
    _auto_column_width(ws, [16, 10, 10, 40, 40])

    # 2) 조직도 시트
    ws = wb.create_sheet(title='조직도')
    ws.merge_cells('A1:D1')
    ws['A1'].value = '조직도 (수행 인력)'
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36
    headers = ['직무', '인원(명)', '등급', '비고']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGNMENT
        c.border = _THIN_BORDER
    row = 4
    total_head = 0
    for org in est.organization:
        values = [org.role, org.headcount, org.grade or '-', org.notes or '-']
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT if col != 2 else \
                Alignment(horizontal='center', vertical='center')
        total_head += org.headcount
        row += 1
    if est.organization:
        # 합계 행
        ws.cell(row=row, column=1, value='합계').font = Font(
            name='맑은 고딕', bold=True, size=11)
        ws.cell(row=row, column=2, value=total_head).font = Font(
            name='맑은 고딕', bold=True, size=11)
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal='center', vertical='center')
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = _THIN_BORDER
    _auto_column_width(ws, [24, 10, 10, 24])

    # 3) M/M 시트
    ws = wb.create_sheet(title='M_M_산정')
    ws.merge_cells('A1:E1')
    ws['A1'].value = (
        f'Man-Month 산정 (총 {est.total_mm:.1f} M/M, '
        f'{est.project_months:.0f}개월 기준)'
    )
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36
    headers = ['직무', '인원(명)', '기간(월)', 'M/M', '등급']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGNMENT
        c.border = _THIN_BORDER
    row = 4
    for mm in est.man_months:
        values = [
            mm.role, mm.headcount, mm.months, mm.total_mm, mm.grade or '-',
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT if col not in (2, 3, 4) else \
                Alignment(horizontal='center', vertical='center')
        row += 1
    if est.man_months:
        ws.cell(row=row, column=1, value='합계').font = Font(
            name='맑은 고딕', bold=True, size=11)
        ws.cell(row=row, column=4, value=est.total_mm).font = Font(
            name='맑은 고딕', bold=True, size=11)
        ws.cell(row=row, column=4).alignment = Alignment(
            horizontal='center', vertical='center')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = _THIN_BORDER
    if est.notes:
        ws.cell(row=row + 2, column=1, value=f'산정 근거: {est.notes}').font = \
            Font(name='맑은 고딕', italic=True, size=10, color='666666')
    _auto_column_width(ws, [24, 10, 10, 10, 10])


def _build_agent_trace_sheet(wb: Workbook, result: RfpAnalysisResult) -> None:
    ws = wb.create_sheet(title='에이전트로그')
    ws.merge_cells('A1:G1')
    ws['A1'].value = '에이전트 호출 트레이스'
    ws['A1'].font = _TITLE_FONT
    ws.row_dimensions[1].height = 36

    headers = ['에이전트', '모델', '소요(ms)', '입력', '출력', '성공', '비고']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGNMENT
        c.border = _THIN_BORDER

    row = 4
    for call in result.agent_trace:
        values = [
            call.get('agent', ''),
            call.get('model', ''),
            call.get('duration_ms', 0),
            call.get('input_chars', 0),
            call.get('output_chars', 0),
            'OK' if call.get('success') else 'FAIL',
            call.get('stop_reason', '') or call.get('error', '')[:80],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT
        row += 1

    _auto_column_width(ws, [20, 22, 10, 10, 10, 8, 60])


# ── Markdown 보고서 ─────────────────────────────────────────

def to_markdown(result: RfpAnalysisResult, filename: str = '') -> str:
    """v2 결과 → 사람 읽기 좋은 한 페이지 마크다운 보고서."""
    lines: list[str] = []

    title = filename or (result.overview.project_name if result.overview else 'RFP 분석 결과')
    lines.append(f'# {title}')
    lines.append('')
    lines.append(f'*생성: {datetime.now().strftime("%Y-%m-%d %H:%M")}*')
    lines.append('')

    # ── 검증 요약 ─────────────────────────────────────────
    if result.validation:
        conf = result.validation.confidence
        badge = '🟢' if conf >= 0.7 else ('🟡' if conf >= 0.5 else '🔴')
        lines.append(f'## {badge} 검증: 신뢰도 {conf:.2f}')
        if result.validation.issues:
            lines.append('')
            for issue in result.validation.issues[:10]:
                sev = issue.get('severity', 'info')
                emoji = {'error': '❌', 'warning': '⚠️', 'info': 'ℹ️'}.get(sev, '·')
                field = issue.get('field', '')
                msg = issue.get('message', '')
                lines.append(f'- {emoji} **[{field}]** {msg}')
        lines.append('')

    # ── 사업 개요 ─────────────────────────────────────────
    if result.overview:
        ov = result.overview
        lines.append('## 사업 개요')
        lines.append('')
        lines.append('| 항목 | 내용 |')
        lines.append('|---|---|')
        fields = [
            ('사업명', ov.project_name),
            ('발주기관', ov.organization),
            ('사업기간', ov.period),
            ('예산', f'{ov.budget} {ov.budget_unit}' if ov.budget else ''),
            ('계약방식', ov.contract_type),
            ('계약유형', ov.contract_method),
            ('참가자격', ov.qualifications),
            ('수행장소', ov.location),
        ]
        for label, value in fields:
            if value:
                lines.append(f'| {label} | {value} |')
        if ov.purpose:
            lines.append('')
            lines.append('**사업 목적**')
            lines.append('')
            lines.append(f'> {ov.purpose}')
        lines.append('')

    # ── 요구사항 ──────────────────────────────────────────
    if result.requirements:
        lines.append(f'## 요구사항 ({len(result.requirements)}건)')
        lines.append('')
        # 카테고리별 그룹
        by_cat: dict[str, list] = {}
        for r in result.requirements:
            key = r.category or r.category_code or '(미분류)'
            by_cat.setdefault(key, []).append(r)
        for cat, reqs in by_cat.items():
            lines.append(f'### {cat} ({len(reqs)}건)')
            lines.append('')
            lines.append('| ID | 명칭 | 요약 |')
            lines.append('|---|---|---|')
            for r in reqs:
                summary = (r.detail or r.definition or r.desc or '').replace('\n', ' ')[:80]
                lines.append(f'| {r.id} | {r.name} | {summary} |')
            lines.append('')

    # ── 배점기준 ──────────────────────────────────────────
    if result.scoring:
        total = sum(s.score for s in result.scoring)
        lines.append(f'## 배점기준 (합계 {total}점)')
        lines.append('')
        lines.append('| 대분류 | 평가항목 | 평가기준 | 배점 |')
        lines.append('|---|---|---|---:|')
        for s in result.scoring:
            criteria = s.criteria.replace('\n', ' ')[:80]
            lines.append(f'| {s.category} | {s.item} | {criteria} | {s.score} |')
        lines.append('')

    # ── 제안목차 ──────────────────────────────────────────
    if result.toc:
        lines.append(f'## 제안목차 (L1 {len(result.toc)}개)')
        lines.append('')
        for t in result.toc:
            lines.append(f'- **{t.number} {t.title}**')
            for child in t.children:
                lines.append(f'  - {child.number} {child.title}')
                for grand in child.children:
                    lines.append(f'    - {grand.number} {grand.title}')
        lines.append('')

    # ── 영업 견적 (WBS / 조직도 / M/M) ────────────────────
    if result.estimation:
        est = result.estimation
        lines.append('## 💼 영업 견적 요약')
        lines.append('')
        lines.append(
            f'**사업기간** {est.project_months:.0f}개월 · '
            f'**총 M/M** {est.total_mm:.1f} · '
            f'**직무** {len(est.organization)}종'
        )
        lines.append('')

        if est.organization:
            lines.append('### 조직도')
            lines.append('')
            lines.append('| 직무 | 인원 | 등급 |')
            lines.append('|---|---:|---|')
            for org in est.organization:
                lines.append(f'| {org.role} | {org.headcount} | {org.grade or "-"} |')
            lines.append('')

        if est.man_months:
            lines.append('### Man-Month 산정')
            lines.append('')
            lines.append('| 직무 | 인원 | 기간(월) | M/M |')
            lines.append('|---|---:|---:|---:|')
            for mm in est.man_months:
                lines.append(
                    f'| {mm.role} | {mm.headcount} | {mm.months:.0f} | '
                    f'{mm.total_mm:.1f} |'
                )
            lines.append(f'| **합계** |  |  | **{est.total_mm:.1f}** |')
            lines.append('')

        if est.wbs:
            lines.append('### WBS (Work Breakdown Structure)')
            lines.append('')
            for phase in est.wbs:
                wrange = (f'W{phase.week_start}~W{phase.week_end}'
                         if phase.week_start else '-')
                lines.append(
                    f'**{phase.name}** · {wrange} · {phase.duration_weeks}주'
                )
                if phase.activities:
                    for a in phase.activities:
                        lines.append(f'  - 활동: {a}')
                if phase.deliverables:
                    for d in phase.deliverables:
                        lines.append(f'  - 산출물: {d}')
                lines.append('')

        if est.notes:
            lines.append(f'> {est.notes}')
            lines.append('')

    # ── 분류 메타 ─────────────────────────────────────────
    if result.classification:
        cls = result.classification
        lines.append('## 분류 메타 (Classifier)')
        lines.append('')
        lines.append(f'- **발주기관 유형**: {cls.org_type}')
        lines.append(f'- **요구사항 ID 체계**: {cls.req_id_scheme}')
        lines.append(f'- **요구사항 형식**: {cls.req_format}')
        lines.append(f'- **배점 형식**: {cls.scoring_format}')
        lines.append(f'- **목차 위치**: {cls.toc_location}')
        lines.append(f'- **문서 규모**: {cls.doc_size_tier}')
        if cls.language_dialect:
            lines.append(f'- **어휘 특이점**: {cls.language_dialect}')
        if cls.notes:
            lines.append('')
            lines.append(f'**메모**: {cls.notes}')
        lines.append('')

    # ── 에이전트 메트릭 ─────────────────────────────────
    if result.agent_trace:
        total_ms = sum(c.get('duration_ms', 0) for c in result.agent_trace)
        total_in = sum(c.get('input_chars', 0) for c in result.agent_trace)
        total_out = sum(c.get('output_chars', 0) for c in result.agent_trace)
        lines.append('---')
        lines.append(f'*에이전트 호출 {len(result.agent_trace)}건, 소요 {total_ms / 1000:.1f}초, '
                     f'입력 {total_in:,}자, 출력 {total_out:,}자*')

    return '\n'.join(lines)
