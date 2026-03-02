"""
RFP 분석 결과 Excel 내보내기 모듈 (Ticket #20)

분석 결과 dict를 4개 시트(사업개요, 요구사항, 배점기준, 제안목차)로
구성된 Excel 워크북으로 변환합니다.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── 스타일 상수 ──────────────────────────────────────────────

_HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
_HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

_LABEL_FONT = Font(name='맑은 고딕', bold=True, size=11, color='2B579A')
_LABEL_FILL = PatternFill(start_color='E8EEF7', end_color='E8EEF7', fill_type='solid')

_BODY_FONT = Font(name='맑은 고딕', size=11)
_BODY_ALIGNMENT = Alignment(vertical='center', wrap_text=True)

_THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

_TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14, color='2B579A')


# ── Public API ───────────────────────────────────────────────

def export_to_excel(data: dict) -> io.BytesIO:
    """분석 결과 dict를 Excel 바이트 스트림으로 변환합니다.

    Args:
        data: RfpAnalysisResult.to_dict() 형식의 dict.
              키: overview, requirements, scoring, toc

    Returns:
        io.BytesIO: xlsx 파일 바이트 스트림 (seek(0) 완료)
    """
    wb = Workbook()

    # 기본 시트 제거 후 순서대로 생성
    wb.remove(wb.active)

    requirements = data.get('requirements', [])
    _build_overview_sheet(wb, data.get('overview'))
    _build_requirements_sheet(wb, requirements)
    _build_scoring_sheet(wb, data.get('scoring', []))
    _build_toc_sheet(wb, data.get('toc', []), requirements)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 사업개요 시트 ────────────────────────────────────────────

_OVERVIEW_LABELS = [
    ('project_name',    '사업명'),
    ('organization',    '발주기관'),
    ('purpose',         '사업 목적'),
    ('period',          '사업기간'),
    ('budget',          '예산'),
    ('contract_type',   '계약 방식'),
    ('contract_method', '계약 유형'),
    ('qualifications',  '참가자격'),
    ('location',        '수행장소'),
]


def _build_overview_sheet(wb: Workbook, overview: dict | None) -> None:
    ws = wb.create_sheet(title='사업개요')

    # 시트 제목
    ws.merge_cells('A1:B1')
    title_cell = ws['A1']
    title_cell.value = '사업개요'
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    if not overview:
        ws['A3'].value = '데이터가 없습니다.'
        ws['A3'].font = _BODY_FONT
        _auto_column_width(ws, [40, 80])
        return

    row = 3
    for key, label in _OVERVIEW_LABELS:
        value = overview.get(key, '')

        # 예산은 단위 포함
        if key == 'budget' and value:
            unit = overview.get('budget_unit', '백만원')
            value = f'{value} {unit}'

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _LABEL_FONT
        label_cell.fill = _LABEL_FILL
        label_cell.border = _THIN_BORDER
        label_cell.alignment = Alignment(horizontal='left', vertical='center')

        value_cell = ws.cell(row=row, column=2, value=value or '-')
        value_cell.font = _BODY_FONT
        value_cell.border = _THIN_BORDER
        value_cell.alignment = _BODY_ALIGNMENT

        ws.row_dimensions[row].height = 28
        row += 1

    _auto_column_width(ws, [20, 80])


# ── 요구사항 시트 ────────────────────────────────────────────

_REQ_HEADERS = ['ID', '분류', '요구사항명', '설명', '필수여부']


def _build_requirements_sheet(wb: Workbook, requirements: list[dict]) -> None:
    ws = wb.create_sheet(title='요구사항')

    # 시트 제목
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = '요구사항 목록'
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # 헤더 행
    for col_idx, header in enumerate(_REQ_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 30

    if not requirements:
        ws.merge_cells('A4:E4')
        ws['A4'].value = '추출된 요구사항이 없습니다.'
        ws['A4'].font = _BODY_FONT
        _auto_column_width(ws, [12, 10, 30, 60, 10])
        return

    row = 4
    for req in requirements:
        values = [
            req.get('id', ''),
            req.get('category', ''),
            req.get('name', ''),
            req.get('desc', ''),
            req.get('level', ''),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT
        row += 1

    _auto_column_width(ws, [12, 10, 30, 60, 10])
    ws.auto_filter.ref = f'A3:E{row - 1}'


# ── 배점기준 시트 ────────────────────────────────────────────

_SCORING_HEADERS = ['대분류', '평가항목', '평가기준', '배점']


def _build_scoring_sheet(wb: Workbook, scoring: list[dict]) -> None:
    ws = wb.create_sheet(title='배점기준')

    # 시트 제목
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = '배점기준'
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # 헤더 행
    for col_idx, header in enumerate(_SCORING_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 30

    if not scoring:
        ws.merge_cells('A4:D4')
        ws['A4'].value = '추출된 배점기준이 없습니다.'
        ws['A4'].font = _BODY_FONT
        _auto_column_width(ws, [15, 25, 60, 10])
        return

    row = 4
    total_score = 0
    for item in scoring:
        score = item.get('score', 0)
        total_score += score
        values = [
            item.get('category', ''),
            item.get('item', ''),
            item.get('criteria', ''),
            score,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _BODY_ALIGNMENT
            if col_idx == 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    # 합계 행
    ws.merge_cells(f'A{row}:C{row}')
    total_label = ws.cell(row=row, column=1, value='합계')
    total_label.font = Font(name='맑은 고딕', bold=True, size=11)
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    total_label.border = _THIN_BORDER
    for c in range(2, 4):
        ws.cell(row=row, column=c).border = _THIN_BORDER

    total_cell = ws.cell(row=row, column=4, value=total_score)
    total_cell.font = Font(name='맑은 고딕', bold=True, size=11)
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_cell.border = _THIN_BORDER

    _auto_column_width(ws, [15, 25, 60, 10])
    ws.auto_filter.ref = f'A3:D{row - 1}'


# ── 제안목차 시트 ────────────────────────────────────────────

_TOC_HEADERS = ['목차', '요구사항 ID', '작성담당자', '설명']

_ROMAN_NUMERALS = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ', 'Ⅷ', 'Ⅸ', 'Ⅹ',
                   'Ⅺ', 'Ⅻ', 'XIII', 'XIV', 'XV']

_L1_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')


def _build_toc_sheet(
    wb: Workbook, toc: list[dict], requirements: list[dict] | None = None,
) -> None:
    ws = wb.create_sheet(title='제안목차')

    # 시트 제목
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = '제안목차'
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # 헤더 행
    for col_idx, header in enumerate(_TOC_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 30

    if not toc:
        ws.merge_cells('A4:D4')
        ws['A4'].value = '추출된 목차가 없습니다.'
        ws['A4'].font = _BODY_FONT
        _auto_column_width(ws, [60, 16, 14, 50])
        return

    rows: list[dict] = []
    _flatten_toc_hierarchical(toc, rows)

    # 요구사항 매핑
    req_map = _match_requirements_to_toc(rows, requirements or [])

    row = 4
    for i, item in enumerate(rows):
        level = item['level']
        label = item['label']
        req_ids = req_map.get(i, '')
        description = item['description']

        # ── 목차 컬럼 (계층형 번호 + 제목, 들여쓰기) ──
        toc_cell = ws.cell(row=row, column=1, value=label)
        if level == 1:
            toc_cell.font = Font(name='맑은 고딕', bold=True, size=11)
            toc_cell.fill = _L1_FILL
        elif level == 2:
            toc_cell.font = Font(name='맑은 고딕', bold=True, size=11)
        else:
            toc_cell.font = _BODY_FONT
        toc_cell.border = _THIN_BORDER
        toc_cell.alignment = Alignment(vertical='center', wrap_text=True)

        # ── 요구사항 ID ──
        req_cell = ws.cell(row=row, column=2, value=req_ids)
        req_cell.font = _BODY_FONT
        req_cell.border = _THIN_BORDER
        req_cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
        if level == 1:
            req_cell.fill = _L1_FILL

        # ── 작성담당자 (빈칸) ──
        author_cell = ws.cell(row=row, column=3, value='')
        author_cell.font = _BODY_FONT
        author_cell.border = _THIN_BORDER
        author_cell.alignment = Alignment(horizontal='center', vertical='center')
        if level == 1:
            author_cell.fill = _L1_FILL

        # ── 설명 ──
        desc_cell = ws.cell(row=row, column=4, value=description)
        desc_cell.font = _BODY_FONT
        desc_cell.border = _THIN_BORDER
        desc_cell.alignment = _BODY_ALIGNMENT
        if level == 1:
            desc_cell.fill = _L1_FILL

        row += 1

    _auto_column_width(ws, [60, 16, 14, 50])


def _flatten_toc_hierarchical(
    items: list[dict],
    out: list[dict],
    parent_prefix: str = '',
    depth: int = 1,
) -> None:
    """재귀적으로 TOC 항목을 계층형 번호 체계로 평탄화합니다.

    번호 체계:
      L1: Ⅰ. Ⅱ. Ⅲ. ...
      L2: 1. 2. 3. ...
      L3: 1.1. 1.2. 2.1. ...
      L4: 1.1.1. 1.1.2. ...
    """
    for idx, item in enumerate(items):
        title = item.get('title', '')
        description = item.get('description', '')

        # ── 계층형 번호 생성 ──
        if depth == 1:
            number = _ROMAN_NUMERALS[idx] if idx < len(_ROMAN_NUMERALS) else str(idx + 1)
            label_number = f'{number}.'
            child_prefix = str(idx + 1)
        elif depth == 2:
            number = str(idx + 1)
            label_number = f'{number}.'
            child_prefix = f'{parent_prefix}{number}.' if parent_prefix else f'{number}.'
        else:
            number = f'{parent_prefix}{idx + 1}.'
            label_number = number
            child_prefix = number

        # ── 들여쓰기 (L2부터 2칸씩) ──
        indent = '  ' * (depth - 1)
        label = f'{indent}{label_number} {title}'

        out.append({
            'level': depth,
            'label': label,
            'title': title,
            'description': description,
        })

        children = item.get('children', [])
        if children:
            _flatten_toc_hierarchical(children, out, child_prefix, depth + 1)


# ── 요구사항 ↔ 목차 매칭 ──────────────────────────────────────

_STOPWORDS = frozenset({
    '사업', '시스템', '구축', '개발', '운영', '관리', '서비스',
    '기능', '구현', '설계', '요구', '사항', '정보', '데이터',
    '처리', '제공', '수행', '지원', '환경', '방안', '현황',
    '개요', '목적', '범위', '기간', '내용', '계획', '추진',
    '대상', '결과', '방법', '작성', '검토', '기타', '기본',
})

_SECTION_TO_CODE: dict[str, list[str]] = {
    '기능': ['SFR'],
    '성능': ['PER'],
    '인터페이스': ['SIR'],
    '보안': ['SER', 'SSR'],
    '데이터': ['DAR', 'SDR'],
    '테스트': ['TER'],
    '품질': ['QUR'],
    '유지관리': ['MPR'],
    '운영': ['MPR', 'PSR'],
    '프로젝트': ['PMR', 'PSR'],
    '장비': ['ECR'],
    '제약': ['COR'],
    '컨설팅': ['CSR'],
}


def _extract_keywords(text: str, use_stopwords: bool = True) -> set[str]:
    """텍스트에서 매칭용 키워드를 추출합니다 (2자 이상 한글 단어)."""
    import re
    if not text:
        return set()
    words = re.findall(r'[가-힣]{2,}', text)
    if use_stopwords:
        return {w for w in words if len(w) >= 2 and w not in _STOPWORDS}
    return {w for w in words if len(w) >= 2}


def _match_requirements_to_toc(
    toc_rows: list[dict], requirements: list[dict],
) -> dict[int, str]:
    """각 TOC 행에 매칭되는 요구사항 ID를 반환합니다.

    매칭 전략:
      1) 키워드 기반: name + definition + detail에서 추출한 키워드와
         목차 title + description의 키워드 겹침으로 점수 산출
      2) 분류 코드 기반: 목차 제목에 분류 키워드가 포함되면 해당 코드의
         요구사항을 보완적으로 매핑

    L1 항목(Ⅰ, Ⅱ 등)은 매핑을 건너뜁니다.

    Returns:
        {행_인덱스: "SFR-001, SFR-002", ...}
    """
    if not requirements:
        return {}

    # ── 요구사항별 키워드 사전 구축 ──
    req_entries: list[tuple[str, set[str], str]] = []  # (id, keywords, code)
    for req in requirements:
        rid = req.get('id', '')
        if not rid:
            continue
        kw = set()
        kw |= _extract_keywords(req.get('name', ''))
        kw |= _extract_keywords(req.get('definition', ''))
        detail = (req.get('detail') or '')[:200]
        kw |= _extract_keywords(detail)
        code = req.get('category_code', '')
        req_entries.append((rid, kw, code))

    if not req_entries:
        return {}

    result: dict[int, str] = {}

    for row_idx, toc_item in enumerate(toc_rows):
        # L1 (대분류: Ⅰ, Ⅱ) 스킵
        if toc_item.get('level', 1) <= 1:
            continue

        title = toc_item.get('title', '')
        if not title:
            continue

        toc_kw = _extract_keywords(title)
        toc_kw |= _extract_keywords(toc_item.get('description', ''))

        if not toc_kw:
            continue

        # ── 방법 1: 키워드 겹침 점수 ──
        scored: list[tuple[str, float, int]] = []
        for rid, req_kw, _ in req_entries:
            if not req_kw:
                continue
            overlap = toc_kw & req_kw
            if not overlap:
                continue
            score = len(overlap) / min(len(toc_kw), len(req_kw))
            if score >= 0.3 or len(overlap) >= 2:
                scored.append((rid, score, len(overlap)))

        # 점수 높은 순, 최대 5개
        scored.sort(key=lambda x: (-x[1], -x[2]))
        matched_ids: list[str] = [s[0] for s in scored[:5]]

        # ── 방법 2: 분류 코드 보완 (키워드 매칭이 부족할 때) ──
        if len(matched_ids) < 2:
            codes: list[str] = []
            for keyword, code_list in _SECTION_TO_CODE.items():
                if keyword in title:
                    codes.extend(code_list)
            if codes:
                for rid, _, rcode in req_entries:
                    if rcode in codes and rid not in matched_ids:
                        matched_ids.append(rid)
                        if len(matched_ids) >= 10:
                            break

        if matched_ids:
            result[row_idx] = ', '.join(dict.fromkeys(matched_ids))

    return result


# ── 유틸리티 ─────────────────────────────────────────────────

def _auto_column_width(ws, min_widths: list[int]) -> None:
    """열 너비를 최소값 기준으로 설정합니다."""
    for col_idx, width in enumerate(min_widths, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def generate_filename(overview: dict | None = None) -> str:
    """다운로드용 파일명을 생성합니다."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if overview and overview.get('project_name'):
        # 파일명에 사용할 수 없는 문자 제거
        safe_name = overview['project_name']
        for ch in r'\/:*?"<>|':
            safe_name = safe_name.replace(ch, '')
        safe_name = safe_name.strip()[:50]
        if safe_name:
            return f'RFP분석_{safe_name}_{timestamp}.xlsx'
    return f'RFP_분석결과_{timestamp}.xlsx'
