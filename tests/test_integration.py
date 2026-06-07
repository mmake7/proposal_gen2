"""통합 테스트 (v2 경로): 업로드 → 분석 → Excel 다운로드 전체 흐름.

v2 파이프라인(ingest/analyze)은 모킹하고, 라우트 연동 + 캐시 + Excel 내보내기
(services.excel_exporter)가 v2 결과 dict로 정상 동작하는지 검증한다.
"""

import io

import pytest
from unittest.mock import patch, MagicMock
from openpyxl import load_workbook


@pytest.mark.integration
class TestUploadToExcel:
    @patch('app.analyze_rfp_document')
    @patch('app.v2_ingest_bytes')
    def test_upload_then_download_excel(self, MockIngest, MockAnalyze, client):
        """v2 분석 후 /api/download/excel로 4시트 Excel(배점 합계 자동계산)을 받는다."""
        MockIngest.return_value = MagicMock(full_text='RFP 본문')
        ro = MagicMock()
        ro.to_dict.return_value = {
            'overview': {'project_name': '통합 테스트 사업', 'budget': '100', 'budget_unit': '백만원'},
            'requirements': [{'category': '기능', 'id': 'FR-001', 'name': '기능1', 'desc': '', 'level': '필수'}],
            'scoring': [{'category': '기술', 'item': '항목1', 'criteria': '기준', 'score': 50},
                        {'category': '가격', 'item': '항목2', 'criteria': '기준', 'score': 50}],
            'toc': [{'level': 1, 'title': '사업 이해', 'children': []}],
        }
        MockAnalyze.return_value = ro

        # 1) 분석 (v2)
        parse_resp = client.post(
            '/api/parse',
            data={'file': (io.BytesIO(b'%PDF fake'), 'rfp.pdf')},
            content_type='multipart/form-data',
        )
        assert parse_resp.status_code == 200

        # 2) Excel 다운로드
        excel_resp = client.get('/api/download/excel')
        assert excel_resp.status_code == 200
        assert excel_resp.content_type == \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # 3) 내용 검증
        wb = load_workbook(io.BytesIO(excel_resp.data))
        assert wb.sheetnames == ['사업개요', '요구사항', '배점기준', '제안목차']
        ws_scoring = wb['배점기준']
        assert ws_scoring.cell(row=ws_scoring.max_row, column=4).value == 100

    def test_excel_download_before_parse_returns_404(self, client):
        """분석 전 Excel 다운로드 시 404."""
        resp = client.get('/api/download/excel')
        assert resp.status_code == 404
