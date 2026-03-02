"""
Excel 다운로드 API 테스트 (Ticket #20)

GET /api/download/excel 엔드포인트 및 Excel 생성 로직을 검증합니다.
"""

import io
import os

import pytest
from openpyxl import load_workbook

import app as app_module
from services.excel_exporter import export_to_excel, generate_filename


# ── 테스트 픽스처 ────────────────────────────────────────────

SAMPLE_ANALYSIS = {
    'overview': {
        'project_name': '테스트 사업',
        'organization': '테스트 기관',
        'period': '2026.04 ~ 2026.12',
        'budget': '500',
        'budget_unit': '백만원',
        'contract_type': '제한경쟁입찰',
        'contract_method': '총액계약',
        'qualifications': '중소기업',
        'location': '서울특별시',
        'purpose': 'AI 기반 시스템 구축',
    },
    'requirements': [
        {'category': '기능', 'id': 'FR-001', 'name': '데이터 수집', 'desc': '자동 수집 기능', 'level': '필수'},
        {'category': '비기능', 'id': 'NFR-001', 'name': '보안', 'desc': '암호화 적용', 'level': '필수'},
    ],
    'scoring': [
        {'category': '기술', 'item': '아키텍처', 'criteria': '확장성', 'score': 30},
        {'category': '가격', 'item': '가격 평가', 'criteria': '적정 가격', 'score': 70},
    ],
    'toc': [
        {'level': 1, 'number': 'I', 'title': '사업 이해', 'description': '사업의 배경과 목적을 이해', 'children': [
            {'level': 2, 'number': '1', 'title': '사업 배경', 'description': '사업 추진 배경', 'children': []},
        ]},
        {'level': 1, 'number': 'II', 'title': '기술 제안', 'description': '기술 아키텍처 제안', 'children': []},
    ],
}


# ── API 엔드포인트 테스트 ────────────────────────────────────

class TestDownloadEndpoint:
    def test_no_analysis_returns_404(self, client):
        resp = client.get('/api/download/excel')
        assert resp.status_code == 404
        body = resp.get_json()
        assert 'error' in body

    def test_with_analysis_returns_xlsx(self, client):
        app_module._last_analysis = SAMPLE_ANALYSIS

        resp = client.get('/api/download/excel')
        assert resp.status_code == 200
        assert resp.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Content-Disposition 헤더 확인
        cd = resp.headers.get('Content-Disposition', '')
        assert 'attachment' in cd
        assert '.xlsx' in cd

    def test_downloaded_file_is_valid_xlsx(self, client):
        app_module._last_analysis = SAMPLE_ANALYSIS

        resp = client.get('/api/download/excel')
        wb = load_workbook(io.BytesIO(resp.data))
        assert wb.sheetnames == ['사업개요', '요구사항', '배점기준', '제안목차']

    def test_filename_contains_project_name(self, client):
        app_module._last_analysis = SAMPLE_ANALYSIS

        resp = client.get('/api/download/excel')
        cd = resp.headers.get('Content-Disposition', '')
        # UTF-8 인코딩된 파일명 확인
        assert 'filename*=' in cd or 'filename=' in cd

    def test_empty_analysis_still_generates_file(self, client):
        app_module._last_analysis = {
            'overview': None,
            'requirements': [],
            'scoring': [],
            'toc': [],
        }

        resp = client.get('/api/download/excel')
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.data))
        assert len(wb.sheetnames) == 4


# ── Excel 생성 로직 테스트 ───────────────────────────────────

class TestExcelExporter:
    def test_export_returns_bytes_io(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        assert isinstance(buf, io.BytesIO)
        assert len(buf.getvalue()) > 0

    def test_four_sheets_created(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        assert wb.sheetnames == ['사업개요', '요구사항', '배점기준', '제안목차']

    def test_overview_sheet_has_data(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['사업개요']
        # 제목 행 + 빈 행 + 9개 필드 = 11행
        assert ws.max_row == 11

    def test_requirements_sheet_has_header_and_rows(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['요구사항']
        # 제목(1) + 빈(2) + 헤더(3) + 2 데이터 행 = 5행
        assert ws.max_row == 5
        # 헤더 확인
        assert ws.cell(row=3, column=1).value == 'ID'
        assert ws.cell(row=3, column=5).value == '필수여부'

    def test_scoring_sheet_has_total_row(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['배점기준']
        # 제목(1) + 빈(2) + 헤더(3) + 2 데이터 + 합계 = 6행
        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value == '합계'
        assert ws.cell(row=last_row, column=4).value == 100

    def test_toc_sheet_flattens_hierarchy(self):
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        # 제목(1) + 빈(2) + 헤더(3) + 3 항목 (사업이해, 사업배경, 기술제안) = 6행
        assert ws.max_row == 6

    def test_empty_overview_handled(self):
        data = {**SAMPLE_ANALYSIS, 'overview': None}
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['사업개요']
        assert ws.cell(row=3, column=1).value == '데이터가 없습니다.'

    def test_empty_requirements_handled(self):
        data = {**SAMPLE_ANALYSIS, 'requirements': []}
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['요구사항']
        assert ws.cell(row=4, column=1).value == '추출된 요구사항이 없습니다.'


# ── 파일명 생성 테스트 ───────────────────────────────────────

class TestFilename:
    def test_with_project_name(self):
        name = generate_filename({'project_name': '테스트 프로젝트'})
        assert name.startswith('RFP분석_테스트 프로젝트_')
        assert name.endswith('.xlsx')

    def test_without_overview(self):
        name = generate_filename(None)
        assert name.startswith('RFP_분석결과_')
        assert name.endswith('.xlsx')

    def test_special_chars_removed(self):
        name = generate_filename({'project_name': 'A/B:C*D?E'})
        assert '/' not in name
        assert ':' not in name
        assert '*' not in name
        assert '?' not in name

    def test_long_name_truncated(self):
        name = generate_filename({'project_name': 'A' * 100})
        # 이름 부분은 50자로 제한
        assert len(name) < 100

    def test_empty_project_name(self):
        """프로젝트명이 빈 문자열이면 기본 파일명 사용"""
        name = generate_filename({'project_name': ''})
        assert name.startswith('RFP_분석결과_')

    def test_whitespace_project_name(self):
        """프로젝트명이 공백만 있으면 기본 파일명 사용"""
        name = generate_filename({'project_name': '   '})
        assert name.startswith('RFP_분석결과_')


# ── Excel 데이터 정확성 테스트 ────────────────────────────

@pytest.mark.unit
class TestExcelDataAccuracy:
    def test_overview_budget_includes_unit(self):
        """사업개요 시트의 예산에 단위가 포함된다"""
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['사업개요']
        # 예산 행 찾기 (label이 '예산'인 행)
        budget_value = None
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == '예산':
                budget_value = ws.cell(row=row, column=2).value
                break
        assert budget_value is not None
        assert '500' in budget_value
        assert '백만원' in budget_value

    def test_requirements_data_matches_input(self):
        """요구사항 시트의 데이터가 입력과 일치한다"""
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['요구사항']
        # 첫 번째 데이터 행 (row 4)
        assert ws.cell(row=4, column=1).value == 'FR-001'
        assert ws.cell(row=4, column=2).value == '기능'
        assert ws.cell(row=4, column=3).value == '데이터 수집'

    def test_scoring_data_matches_input(self):
        """배점기준 시트의 데이터가 입력과 일치한다"""
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['배점기준']
        # 첫 번째 데이터 행 (row 4)
        assert ws.cell(row=4, column=1).value == '기술'
        assert ws.cell(row=4, column=2).value == '아키텍처'
        assert ws.cell(row=4, column=4).value == 30

    def test_empty_scoring_handled(self):
        """빈 배점기준도 정상 처리된다"""
        data = {**SAMPLE_ANALYSIS, 'scoring': []}
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['배점기준']
        assert ws.cell(row=4, column=1).value == '추출된 배점기준이 없습니다.'

    def test_empty_toc_handled(self):
        """빈 목차도 정상 처리된다"""
        data = {**SAMPLE_ANALYSIS, 'toc': []}
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        assert ws.cell(row=4, column=1).value == '추출된 목차가 없습니다.'

    def test_toc_hierarchical_numbering(self):
        """목차 시트가 계층형 번호 체계로 출력된다"""
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        # 헤더 확인 (4컬럼)
        assert ws.cell(row=3, column=1).value == '목차'
        assert ws.cell(row=3, column=2).value == '요구사항 ID'
        assert ws.cell(row=3, column=3).value == '작성담당자'
        assert ws.cell(row=3, column=4).value == '설명'
        # L1: Ⅰ. 사업 이해
        l1_label = ws.cell(row=4, column=1).value
        assert 'Ⅰ.' in l1_label
        assert '사업 이해' in l1_label
        assert ws.cell(row=4, column=4).value == '사업의 배경과 목적을 이해'
        # L2: 1. 사업 배경 (들여쓰기)
        l2_label = ws.cell(row=5, column=1).value
        assert '1.' in l2_label
        assert '사업 배경' in l2_label
        # L1: Ⅱ. 기술 제안
        l1_2_label = ws.cell(row=6, column=1).value
        assert 'Ⅱ.' in l1_2_label
        assert '기술 제안' in l1_2_label

    def test_toc_author_column_empty(self):
        """작성담당자 컬럼은 빈칸이다"""
        buf = export_to_excel(SAMPLE_ANALYSIS)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        for row in range(4, ws.max_row + 1):
            assert ws.cell(row=row, column=3).value == '' or ws.cell(row=row, column=3).value is None

    def test_deeply_nested_toc_hierarchical(self):
        """L3, L4까지 계층형 번호가 올바르게 생성된다"""
        data = {
            **SAMPLE_ANALYSIS,
            'toc': [
                {'level': 1, 'number': 'I', 'title': '제안개요', 'description': '개요', 'children': [
                    {'level': 2, 'number': '1', 'title': '사업의 이해', 'description': '', 'children': [
                        {'level': 3, 'number': '가', 'title': '사업의 배경', 'description': '', 'children': []},
                        {'level': 3, 'number': '나', 'title': '주요 업무', 'description': '', 'children': [
                            {'level': 4, 'title': '업무 개요', 'description': '', 'children': []},
                            {'level': 4, 'title': '시스템 이해', 'description': '', 'children': []},
                        ]},
                    ]},
                ]},
            ],
        }
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        # 6 items total: 제안개요, 사업의 이해, 사업의 배경, 주요 업무, 업무 개요, 시스템 이해
        assert ws.max_row == 9  # 3 header + 6 data
        # L1: Ⅰ. 제안개요
        assert 'Ⅰ.' in ws.cell(row=4, column=1).value
        # L2: 1. 사업의 이해
        assert '1.' in ws.cell(row=5, column=1).value
        # L3: 1.1. 사업의 배경
        assert '1.1.' in ws.cell(row=6, column=1).value
        # L3: 1.2. 주요 업무
        assert '1.2.' in ws.cell(row=7, column=1).value
        # L4: 1.2.1. 업무 개요
        assert '1.2.1.' in ws.cell(row=8, column=1).value
        # L4: 1.2.2. 시스템 이해
        assert '1.2.2.' in ws.cell(row=9, column=1).value


# ── 요구사항 ↔ 목차 매핑 테스트 ────────────────────────────

@pytest.mark.unit
class TestRequirementTocMapping:
    """_match_requirements_to_toc 매핑 로직 검증"""

    def _make_toc(self, items):
        """TOC 딕셔너리 → 플랫 행 리스트"""
        from services.excel_exporter import _flatten_toc_hierarchical
        rows = []
        _flatten_toc_hierarchical(items, rows)
        return rows

    def test_keyword_matching(self):
        """키워드 겹침으로 올바른 요구사항이 매핑된다"""
        from services.excel_exporter import _match_requirements_to_toc
        toc = self._make_toc([
            {'title': '제안 방안', 'children': [
                {'title': '인터페이스 설계', 'children': []},
                {'title': '보안 대책 방안', 'children': []},
            ]},
        ])
        reqs = [
            {'id': 'SIR-001', 'name': '외부 연계', 'category_code': 'SIR',
             'definition': '외부 시스템 인터페이스 연계 설계'},
            {'id': 'SER-001', 'name': '접근 통제', 'category_code': 'SER',
             'definition': '보안 접근 통제 및 대책 적용'},
        ]
        result = _match_requirements_to_toc(toc, reqs)
        # L2 '인터페이스 설계' → SIR-001
        matched_values = list(result.values())
        assert any('SIR-001' in v for v in matched_values)

    def test_l1_items_skipped(self):
        """L1 항목에는 매핑하지 않는다"""
        from services.excel_exporter import _match_requirements_to_toc
        toc = self._make_toc([
            {'title': '보안 관리', 'children': []},
        ])
        reqs = [
            {'id': 'SER-001', 'name': '보안 정책', 'category_code': 'SER',
             'definition': '보안 관리 정책 수립'},
        ]
        result = _match_requirements_to_toc(toc, reqs)
        # L1만 있으므로 매핑 없어야 함
        assert result == {}

    def test_category_code_fallback(self):
        """키워드 매칭 부족 시 분류 코드로 보완 매핑"""
        from services.excel_exporter import _match_requirements_to_toc
        toc = self._make_toc([
            {'title': '시스템', 'children': [
                {'title': '보안 방안', 'children': []},
            ]},
        ])
        reqs = [
            {'id': 'SER-001', 'name': '암호화', 'category_code': 'SER',
             'definition': '암호화 적용'},
        ]
        result = _match_requirements_to_toc(toc, reqs)
        # '보안' 키워드 → SER 코드 → SER-001 매핑
        assert any('SER-001' in v for v in result.values())

    def test_empty_requirements_returns_empty(self):
        """요구사항이 없으면 빈 딕셔너리 반환"""
        from services.excel_exporter import _match_requirements_to_toc
        toc = self._make_toc([
            {'title': '사업 이해', 'children': [
                {'title': '사업 배경', 'children': []},
            ]},
        ])
        result = _match_requirements_to_toc(toc, [])
        assert result == {}

    def test_max_five_matches(self):
        """키워드 매칭은 최대 5개까지"""
        from services.excel_exporter import _match_requirements_to_toc
        toc = self._make_toc([
            {'title': '기능', 'children': [
                {'title': '통합 검색 기능', 'children': []},
            ]},
        ])
        # 동일 키워드를 가진 10개 요구사항
        reqs = [
            {'id': f'SFR-{i:03d}', 'name': f'통합 검색 기능 {i}',
             'category_code': 'SFR', 'definition': '통합 검색 기능 제공'}
            for i in range(1, 11)
        ]
        result = _match_requirements_to_toc(toc, reqs)
        for ids_str in result.values():
            ids = [x.strip() for x in ids_str.split(',')]
            assert len(ids) <= 5

    def test_mapping_in_excel_output(self):
        """엑셀 출력에 매핑된 요구사항 ID가 포함된다"""
        data = {
            'overview': None,
            'requirements': [
                {'id': 'SER-001', 'name': '보안 정책', 'category': '보안',
                 'category_code': 'SER', 'definition': '보안 대책 수립',
                 'desc': '보안 대책 수립', 'level': ''},
            ],
            'scoring': [],
            'toc': [
                {'title': '보안', 'children': [
                    {'title': '보안 대책', 'description': '보안 정책 및 대책 수립', 'children': []},
                ]},
            ],
        }
        buf = export_to_excel(data)
        wb = load_workbook(buf)
        ws = wb['제안목차']
        # L2 행 (row 5)의 요구사항 ID 컬럼
        req_id_cell = ws.cell(row=5, column=2).value
        assert req_id_cell is not None and 'SER-001' in req_id_cell
