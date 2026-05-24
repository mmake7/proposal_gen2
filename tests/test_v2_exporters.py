"""v2 결과 → Excel/Markdown 출력 테스트."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from services_v2.results.exporters import to_excel, to_markdown
from services_v2.results.rfp_analysis import (
    Overview, Requirement, RfpAnalysisResult, RfpClassification,
    ScoringItem, TocItem, ValidationReport,
)


def build_sample_result() -> RfpAnalysisResult:
    return RfpAnalysisResult(
        overview=Overview(
            project_name='2026년 통합관제시스템 고도화',
            organization='행정안전부',
            period='2026.04 ~ 2026.12',
            budget='1,500',
            budget_unit='백만원',
            contract_type='제한경쟁입찰',
            contract_method='협상에 의한 계약',
            qualifications='중소SW사업자',
            location='서울',
            purpose='기존 시스템을 고도화하여 안정성·효율을 제고함',
        ),
        requirements=[
            Requirement(id='SFR-001', name='사용자 인증', category='기능 요구사항',
                        category_code='SFR', detail='2FA + 캡차', parse_method='table'),
            Requirement(id='SFR-002', name='대시보드', category='기능 요구사항',
                        category_code='SFR', detail='주요 지표 카드', parse_method='table'),
            Requirement(id='SER-001', name='보안 일반', category='보안 요구사항',
                        category_code='SER', detail='ISMS 준수', parse_method='table'),
        ],
        scoring=[
            ScoringItem(category='정성평가', item='사업이해도',
                       criteria='사업 목적 이해', score=20),
            ScoringItem(category='정성평가', item='기술방안',
                       criteria='구현 방안 구체성', score=30),
            ScoringItem(category='정량평가', item='경영상태',
                       criteria='신용등급', score=50),
        ],
        toc=[
            TocItem(level=1, number='Ⅰ.', title='사업개요',
                   children=[TocItem(level=2, number='1.', title='추진배경')]),
            TocItem(level=1, number='Ⅱ.', title='기술제안'),
        ],
        classification=RfpClassification(
            org_type='중앙부처', req_id_scheme='SFR/SER',
            req_format='table', scoring_format='table',
            toc_location='front', doc_size_tier='medium',
            language_dialect='정의/세부내용 필드명',
            notes='행정안전부 표준 SI 양식',
        ),
        validation=ValidationReport(
            confidence=0.85,
            issues=[
                {'severity': 'info', 'field': 'overview',
                 'message': '주요 필드 모두 채워짐'},
                {'severity': 'warning', 'field': 'requirements',
                 'message': 'PSR 분류 누락'},
            ],
            retry_targets=[],
        ),
        agent_trace=[
            {'agent': 'classifier', 'model': 'claude-opus-4-7',
             'duration_ms': 5000, 'input_chars': 30000, 'output_chars': 400,
             'success': True, 'error': '', 'stop_reason': 'end_turn'},
            {'agent': 'overview', 'model': 'claude-opus-4-7',
             'duration_ms': 4000, 'input_chars': 30000, 'output_chars': 500,
             'success': True, 'error': '', 'stop_reason': 'end_turn'},
            {'agent': 'table_parser', 'model': 'rule-based',
             'duration_ms': 30, 'input_chars': 50000, 'output_chars': 8000,
             'success': True, 'error': '', 'stop_reason': ''},
        ],
    )


# ── Excel ──────────────────────────────────────────────────

class TestExcelExport:
    def test_returns_bytesio(self):
        result = build_sample_result()
        buf = to_excel(result)
        assert isinstance(buf, io.BytesIO)
        assert buf.tell() == 0  # seek(0) 완료

    def test_contains_v1_sheets(self):
        result = build_sample_result()
        buf = to_excel(result)
        wb = load_workbook(buf)
        for sheet in ['사업개요', '요구사항', '배점기준', '제안목차']:
            assert sheet in wb.sheetnames, f'{sheet} 시트 누락'

    def test_contains_v2_meta_sheets(self):
        result = build_sample_result()
        buf = to_excel(result)
        wb = load_workbook(buf)
        for sheet in ['분석메타', '검증결과', '에이전트로그']:
            assert sheet in wb.sheetnames, f'{sheet} 시트 누락'

    def test_classification_values_present(self):
        result = build_sample_result()
        buf = to_excel(result)
        wb = load_workbook(buf)
        ws = wb['분석메타']
        # 모든 셀에서 값 검색
        values = [c.value for row in ws.iter_rows() for c in row]
        assert '중앙부처' in values
        assert '정의/세부내용 필드명' in values

    def test_validation_confidence_in_sheet(self):
        result = build_sample_result()
        buf = to_excel(result)
        wb = load_workbook(buf)
        ws = wb['검증결과']
        values = [c.value for row in ws.iter_rows() for c in row]
        assert '0.85' in values

    def test_agent_trace_rows(self):
        result = build_sample_result()
        buf = to_excel(result)
        wb = load_workbook(buf)
        ws = wb['에이전트로그']
        agents = [ws.cell(row=r, column=1).value for r in range(4, 7)]
        assert 'classifier' in agents
        assert 'table_parser' in agents

    def test_empty_result_does_not_crash(self):
        empty = RfpAnalysisResult()
        buf = to_excel(empty)
        wb = load_workbook(buf)
        assert '사업개요' in wb.sheetnames


# ── Markdown ────────────────────────────────────────────────

class TestMarkdownExport:
    def test_contains_title_and_sections(self):
        result = build_sample_result()
        md = to_markdown(result)
        assert md.startswith('# 2026년 통합관제시스템 고도화')
        assert '## 사업 개요' in md
        assert '## 요구사항 (3건)' in md
        assert '## 배점기준 (합계 100점)' in md
        assert '## 제안목차 (L1 2개)' in md
        assert '## 분류 메타' in md

    def test_validation_badge(self):
        result = build_sample_result()
        md = to_markdown(result)
        # confidence 0.85 → 녹색 배지
        assert '🟢' in md
        assert '0.85' in md

    def test_low_confidence_shows_red(self):
        result = build_sample_result()
        result.validation.confidence = 0.3
        md = to_markdown(result)
        assert '🔴' in md

    def test_requirements_grouped_by_category(self):
        result = build_sample_result()
        md = to_markdown(result)
        # SFR 2건 + SER 1건 — 카테고리별 섹션 헤더
        assert '### 기능 요구사항 (2건)' in md
        assert '### 보안 요구사항 (1건)' in md

    def test_toc_tree_indentation(self):
        result = build_sample_result()
        md = to_markdown(result)
        # L1 + L2 들여쓰기
        assert '- **Ⅰ. 사업개요**' in md
        assert '  - 1. 추진배경' in md

    def test_purpose_as_blockquote(self):
        result = build_sample_result()
        md = to_markdown(result)
        assert '> 기존 시스템을 고도화' in md

    def test_agent_metrics_footer(self):
        result = build_sample_result()
        md = to_markdown(result)
        assert '에이전트 호출 3건' in md

    def test_empty_result_does_not_crash(self):
        md = to_markdown(RfpAnalysisResult())
        # 비어도 최소 타이틀
        assert md.startswith('# RFP 분석 결과')
